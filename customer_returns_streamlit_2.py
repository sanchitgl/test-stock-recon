import pandas as pd
import os
import sys
import datetime
import openpyxl
import numpy as np
import recordlinkage
# import dateparser
# import dateutil

def reconcile(payment_report, returns_report, reimbursement_report, inventory_ledger):
	data_to_excel = pd.ExcelWriter('temp/customer_returns_reco.xlsx')

	#--------------------------------------Step 1------------------------------------------------
	payment = payment_report
	payment_refund = payment[payment['type'] == 'Refund']
	payment_refund['month'] = payment_refund['date/time'].str[0:3]
	payment_refund = payment_refund[payment_refund['month'] == 'Feb']
	# payment_refund['date/time'] = pd.to_datetime(payment_refund['date/time'].apply(dateparser.parse), utc=True)
	payment_refund['sku'].fillna('Not Available', inplace=True)
	payment_refund['sku'] = payment_refund['sku'].astype(str)
	payment_refund['sku'] = payment_refund['sku'].str.replace('_New', '')
	payment_refund['sku'] = payment_refund['sku'].str.replace('_NEW', '')
	payment_refund['sku'] = np.where(payment_refund['sku'].str.len() > 13, payment_refund['sku'].str[8:20], payment_refund['sku'])
	payment_refund['total'] = -payment_refund['total'].astype(float)
	payment_refund = payment_refund[['date/time', 'order id', 'sku', 'quantity', 'total']].rename(columns={'order id': 'order-id', 'quantity': 'quantity-refund'})
	payment_refund = payment_refund.groupby(['order-id', 'sku']).agg({'quantity-refund': 'sum', 'total': 'sum'}).reset_index()
	print(payment_refund)
	# sys.exit()

	customer_returns = returns_report
	customer_returns = customer_returns.rename(columns={'quantity': 'quantity-returns'})
	customer_returns['sku'].fillna('Not Available', inplace=True)
	customer_returns['sku'] = customer_returns['sku'].astype(str)
	customer_returns['sku'] = customer_returns['sku'].str.replace('_New', '')
	customer_returns['sku'] = customer_returns['sku'].str.replace('_NEW', '')
	customer_returns['sku'] = np.where(customer_returns['sku'].str.len() > 13, customer_returns['sku'].str[8:20], customer_returns['sku'])
	customer_returns_grouped = customer_returns.groupby(['order-id', 'sku', 'status']).agg({'quantity-returns': 'sum'}).reset_index()

	payment_refund_merged = payment_refund.merge(customer_returns_grouped, on=['order-id', 'sku'], how='left')
	print("#######")
	print(payment_refund_merged)
	payment_refund_merged['status'].fillna('Not found in customer returns', inplace=True)
	## New way
	payment_refund_merged['quantity-returns'].fillna(0, inplace=True)
	payment_refund_merged['Reimbursed'] = np.where(payment_refund_merged['status'] == 'Reimbursed', payment_refund_merged['quantity-returns'], 0)
	payment_refund_merged['Returned to Inventory'] = np.where((payment_refund_merged['status'] == 'Unit returned to inventory') | (payment_refund_merged['status'] == 'Repackaged Successfully'), payment_refund_merged['quantity-returns'], 0)
	payment_refund_merged['Quantity Difference'] = payment_refund_merged['quantity-refund'] - payment_refund_merged['quantity-returns']
	# payment_refund_merged.to_excel(data_to_excel, sheet_name='1. Refunds vs Returns')
	# data_to_excel.save()
	# sys.exit()
	payment_refund_merged['Amount Difference'] = payment_refund_merged['total'] / payment_refund_merged['quantity-refund'] * payment_refund_merged['Quantity Difference']

	
	payment_refund_merged = payment_refund_merged.groupby(['order-id', 'sku']).agg({'quantity-refund': 'sum', 'total': 'sum', 'quantity-returns': 'sum', 'Reimbursed': 'sum', 'Returned to Inventory': 'sum', 'Quantity Difference': 'sum', 'Amount Difference': 'sum'})
	# payment_refund_merged['quantity-difference'] = payment_refund_merged['quantity-refund'] - payment_refund_merged['quantity-returns']
	# payment_refund_merged['amount-difference'] = payment_refund_merged['total'] / payment_refund_merged['quantity-refund'] * payment_refund_merged['quantity-difference']
	# payment_refund_merged.set_index(['order-id', 'sku', 'quantity-refund', 'total' 'ful'], inplace=True)
	print("------------------------")
	print(payment_refund_merged)

	payment_refund_merged = payment_refund_merged.reset_index()
	payment_refund_merged_renamed = payment_refund_merged.rename(columns={'order-id': 'Order ID', 'sku': 'SKU', 'quantity-refund': 'Refund Quantity', 'total': 'Refund Amount', 'quantity-returns': 'Customer Returns', 'Returned to Inventory': 'Added to Inventory'})
	payment_refund_merged_renamed.to_excel(data_to_excel, sheet_name='1. Refunds vs Returns', index=False)
	# data_to_excel.save()
	# sys.exit()

	#-------------------------------------------Step 2-----------------------------------------
	customer_returns_grouped = customer_returns.groupby(['order-id', 'sku', 'status', 'fulfillment-center-id']).agg({'quantity-returns': 'sum'}).reset_index()

	payment_refund_merged = payment_refund.merge(customer_returns_grouped, on=['order-id', 'sku'], how='left')
	payment_refund_merged['status'].fillna('Not found in customer returns', inplace=True)
	payment_refund_merged['quantity-returns'].fillna(0, inplace=True)
	payment_refund_merged['Reimbursed'] = np.where(payment_refund_merged['status'] == 'Reimbursed', payment_refund_merged['quantity-returns'], 0)
	# payment_refund_merged = payment_refund_merged.reset_index()
	payment_reimbursed = payment_refund_merged[payment_refund_merged['Reimbursed'] != 0]
	# payment_inventory = payment_refund[(payment_refund['status'] == 'Unit returned to inventory') | (payment_refund['status'] == 'Repackaged Successfully')]
	# payment_notfound = payment_refund[payment_refund['status'] == 'Not found in customer returns']
	# # payment_refund.to_csv('payment_refund.csv')

	# print(payment_reimbursed)
	# print(payment_inventory)
	# print(payment_notfound)
	# sys.exit()

	reimbursement = reimbursement_report
	reimbursement = reimbursement[reimbursement['reason'] == 'CustomerReturn']
	reimbursement['sku'].fillna('Not Available', inplace=True)
	reimbursement['sku'] = reimbursement['sku'].astype(str)
	reimbursement['sku'] = reimbursement['sku'].str.replace('_New', '')
	reimbursement['sku'] = reimbursement['sku'].str.replace('_NEW', '')
	reimbursement['sku'] = np.where(reimbursement['sku'].str.len() > 13, reimbursement['sku'].str[8:20], reimbursement['sku'])
	reimbursement = reimbursement[['approval-date', 'amazon-order-id', 'sku', 'amount-total', 'quantity-reimbursed-cash', 'quantity-reimbursed-inventory', 'quantity-reimbursed-total']].rename(columns={'amazon-order-id': 'order-id'})
	reimbursement_grouped = reimbursement.groupby(['order-id', 'sku']).agg({'amount-total': 'sum', 'quantity-reimbursed-cash': 'sum', 'quantity-reimbursed-inventory': 'sum', 'quantity-reimbursed-total': 'sum'})
	# reimbursement.to_csv('reimbursement.csv')


	payment_reimbursed_grouped = payment_reimbursed.merge(reimbursement_grouped, on=['order-id', 'sku'], how='left')


	# payment_reimbursed_grouped.drop(['status', 'Quantity Difference', 'Amount Difference'], axis=1, inplace=True)
	payment_reimbursed_grouped['amount-total'].fillna(0, inplace=True)
	payment_reimbursed_grouped['quantity-reimbursed-cash'].fillna(0, inplace=True)
	payment_reimbursed_grouped['quantity-reimbursed-inventory'].fillna(0, inplace=True)
	payment_reimbursed_grouped['quantity-reimbursed-total'].fillna(0, inplace=True)
	payment_reimbursed_grouped['quantity-difference'] = payment_reimbursed_grouped['quantity-returns'] - payment_reimbursed_grouped['quantity-reimbursed-total']
	payment_reimbursed_grouped['amount-difference'] = payment_reimbursed_grouped['total'] / payment_reimbursed_grouped['Reimbursed'] * payment_reimbursed_grouped['quantity-difference']
	payment_reimbursed_grouped.drop('quantity-refund', axis=1, inplace=True)
	# payment_reimbursed.set_index([''])
	# print(payment_reimbursed.info())

	payment_reimbursed_grouped_to_excel = payment_reimbursed_grouped.drop(['status', 'fulfillment-center-id', 'quantity-returns'], axis=1) 
	payment_reimbursed_grouped_to_excel = payment_reimbursed_grouped_to_excel.rename(columns={'order-id': 'Order ID', 'sku': 'SKU', 'total': 'Refund Amount', 'Reimbursed': 'Reimbursed Quantity', 'amount-total': 'Reimbursed Amount', 'quantity-reimbursed-cash': 'Cash Reimbursement', 'quantity-reimbursed-inventory': 'Inventory Reimbursement', 'quantity-reimbursed-total': 'Total Reimbursement', 'quantity-difference': 'Quantity Difference', 'amount-difference': 'Amount Difference'})
	payment_reimbursed_grouped_to_excel.to_excel(data_to_excel, sheet_name='2. Returns vs Reimbursements', index=False)
	# data_to_excel.save()
	# sys.exit()

	#------------------------------------------Step 3------------------------------------------------
	payment_reimbursed_cash = payment_reimbursed_grouped.drop(['quantity-returns', 'quantity-reimbursed-inventory', 'quantity-reimbursed-total', 'quantity-difference'], axis=1)
	payment_reimbursed_cash = payment_reimbursed_cash[payment_reimbursed_cash['quantity-reimbursed-cash'] != 0]

	payment_returns = payment[payment['description'] == 'FBA Inventory Reimbursement - Customer Return']
	payment_returns['sku'].fillna('Not Available', inplace=True)
	payment_returns['sku'] = payment_returns['sku'].astype(str)
	payment_returns['sku'] = payment_returns['sku'].str.replace('_New', '')
	payment_returns['sku'] = payment_returns['sku'].str.replace('_NEW', '')
	payment_returns['sku'] = np.where(payment_returns['sku'].str.len() > 13, payment_returns['sku'].str[8:20], payment_returns['sku'])
	payment_returns['total'] = payment_returns['total'].astype(float)
	payment_returns['quantity'].fillna(0, inplace=True)
	payment_returns['total'].fillna(0, inplace=True)
	payment_returns = payment_returns[['order id', 'sku', 'quantity', 'total']].rename(columns={'order id': 'order-id', 'quantity': 'quantity-settled', 'total': 'amount-settled'})
	payment_returns = payment_returns.groupby(['order-id', 'sku']).agg({'quantity-settled': 'sum', 'amount-settled': 'sum'}).reset_index()
	# print(payment_returns.info())

	payment_reimbursed_cash = payment_reimbursed_cash.merge(payment_returns, on=['order-id', 'sku'], how='left')
	payment_reimbursed_cash['quantity-settled'].fillna(0, inplace=True)
	payment_reimbursed_cash['amount-settled'].fillna(0, inplace=True)
	payment_reimbursed_cash['quantity-difference'] = payment_reimbursed_cash['quantity-reimbursed-cash'] - payment_reimbursed_cash['quantity-settled']
	payment_reimbursed_cash['amount-difference'] = payment_reimbursed_cash['amount-total'] - payment_reimbursed_cash['amount-settled']
	payment_reimbursed_cash = payment_reimbursed_cash[['order-id', 'sku', 'quantity-reimbursed-cash', 'amount-total', 'quantity-settled', 'amount-settled', 'quantity-difference', 'amount-difference']]

	payment_reimbursed_cash = payment_reimbursed_cash.rename(columns={'order-id': 'Order ID', 'sku': 'SKU', 'quantity-reimbursed-cash': 'Quantity Reimbursed', 'amount-total': 'Amount Reimbursed', 'quantity-settled': 'Quantity Settled', 'amount-settled': 'Amount Settled', 'quantity-difference': 'Quantity Difference', 'amount-difference': 'Amount Difference'})	
	payment_reimbursed_cash.to_excel(data_to_excel, sheet_name='3. Cash Reimbursement', index=False)
	# print(payment_reimbursed_cash)

	#-------------------------------------------Step 4--------------------------------------------
	customer_returns_datewise = customer_returns.groupby(['order-id', 'sku', 'status', 'return-date', 'fulfillment-center-id']).agg({'quantity-returns': 'sum'}).reset_index()
	payment_refund_datewise = payment_refund.merge(customer_returns_datewise, on=['order-id', 'sku'], how='left')
	# print(payment_refund_datewise.info())
	# sys.exit()
	payment_refund_datewise['status'].fillna('Not found in customer returns', inplace=True)
	payment_refund_datewise['quantity-returns'].fillna(0, inplace=True)
	payment_inventory_datewise = payment_refund_datewise[(payment_refund_datewise['status'] == 'Unit returned to inventory') | (payment_refund_datewise['status'] == 'Repackaged Successfully')]
	payment_inventory_datewise = payment_inventory_datewise[['sku', 'fulfillment-center-id', 'return-date', 'quantity-returns', 'total']]
	payment_inventory_datewise = payment_inventory_datewise.loc[payment_inventory_datewise.index.repeat(payment_inventory_datewise['quantity-returns'])].reset_index(drop=True)
	payment_inventory_datewise['quantity-returns'] = 1
	payment_inventory_datewise['return-date'] = payment_inventory_datewise['return-date'].str[0:10]
	payment_inventory_datewise['return-date'] = pd.to_datetime(payment_inventory_datewise['return-date'], format='%Y-%m-%d')
	# payment_inventory_datewise = payment_inventory_datewise.groupby(['sku', 'fulfillment-center-id', 'return-date']).agg({'quantity-returns': 'sum'}).reset_index()
	# payment_inventory_datewise.set_index(['order-id', 'sku', 'quantity-refund', 'total', 'fulfillment-center-id', 'status'], inplace=True)
	# payment_inventory_datewise.to_excel('payment_inventory_datewise.xlsx')
	# print(payment_inventory_datewise)
	# sum_ = payment_inventory_datewise['quantity-returns'].sum()
	# print(sum_)
	# sys.exit()

	inventory = inventory_ledger
	inventory = inventory[(inventory['Event Type'] == 'CustomerReturns') & (inventory['Disposition'] == 'SELLABLE')]
	inventory = inventory[['Date', 'MSKU', 'Quantity', 'Fulfillment Center']].rename(columns={'MSKU': 'sku', 'Fulfillment Center': 'fulfillment-center-id'})
	inventory['Date'] = pd.to_datetime(inventory['Date'], format='%m/%d/%Y')
	inventory['sku'] = np.where(inventory['sku'].str[0:4] == 'amzn', inventory['sku'].str[8:20], inventory['sku'])
	inventory['sku'] = inventory['sku'].astype(str).str[0:12]
	inventory = inventory.loc[inventory.index.repeat(inventory['Quantity'])].reset_index(drop=True)
	inventory['Quantity'] = 1


	indexer1 = recordlinkage.Index()
	indexer1.block(left_on=['sku', 'fulfillment-center-id'], right_on=['sku', 'fulfillment-center-id'])
	comparisons1 = indexer1.index(payment_inventory_datewise, inventory)
	compare1 = recordlinkage.Compare()
	compare1.exact('sku', 'sku', label='sku_match')
	compare1.exact('fulfillment-center-id', 'fulfillment-center-id', label='center_match')
	result1 = compare1.compute(comparisons1, payment_inventory_datewise, inventory)
	result_reset1 = result1.reset_index().drop(['sku_match', 'center_match'], axis=1)

	result_reset1 = result_reset1.merge(payment_inventory_datewise, left_on='level_0', right_index=True)
	result_reset1 = result_reset1.merge(inventory[['Date', 'Quantity']], left_on='level_1', right_index=True)
	result_reset1['date_difference'] = result_reset1['Date'] - result_reset1['return-date']
	result_reset1 = result_reset1[result_reset1['date_difference'] >= datetime.timedelta(days=0)]
	result_reset1.sort_values(by=['date_difference', 'level_0'], inplace=True)

	returns1 = pd.DataFrame()
	while len(result_reset1) > 0:
		returns1 = returns1.append(result_reset1.iloc[0])
		result_reset1 = result_reset1[(result_reset1['level_0'] != result_reset1['level_0'].iloc[0]) & (result_reset1['level_1'] != result_reset1['level_1'].iloc[0])]

	# returns.to_excel('returns1.xlsx')
	# print(returns1)

	payment_inventory_datewise = payment_inventory_datewise.merge(returns1[['level_0', 'Date', 'Quantity', 'date_difference']], left_index=True, right_on='level_0', how='left').drop('level_0', axis=1)
	payment_inventory_datewise['quantity-difference'] = np.where(payment_inventory_datewise['Quantity'] == 1, 0, 1)
	payment_inventory_datewise['amount-difference'] = np.where(payment_inventory_datewise['Quantity'] == 1, 0, payment_inventory_datewise['total'])
	# payment_inventory_datewise = payment_inventory_datewise[['sku', 'fulfillment-center-id', 'return-date_x', '']]

	payment_inventory_datewise['return-date'] = payment_inventory_datewise['return-date'].dt.strftime('%m/%d/%Y')
	payment_inventory_datewise['Date'] = payment_inventory_datewise['Date'].dt.strftime('%m/%d/%Y')
	payment_inventory_datewise = payment_inventory_datewise.rename(columns={'sku': 'SKU', 'fulfillment-center-id': 'Fulfillment Center', 'return-date': 'Return Date', 'quantity-returns': 'Return Quantity', 'total': 'Amount', 'Date': 'Date added to Inventory', 'Quantity': 'Quantity Added', 'date_difference': 'Days Difference', 'quantity-difference': 'Quantity Difference', 'amount-difference': 'Amount Difference'})
	payment_inventory_datewise.to_excel(data_to_excel, sheet_name='4. Returned to Inventory', index=False)

	#------------------------------------Step 5-----------------------------------------------------
	reimbursement_datewise = reimbursement.groupby(['order-id', 'sku', 'approval-date']).agg({'amount-total': 'sum', 'quantity-reimbursed-cash': 'sum', 'quantity-reimbursed-inventory': 'sum', 'quantity-reimbursed-total': 'sum'}).reset_index()

	# print(payment_reimbursed_inventory.info())

	payment_reimbursed_datewise = payment_reimbursed.merge(reimbursement_datewise, on=['order-id', 'sku'], how='left')
	payment_reimbursed_datewise.drop('status', axis=1, inplace=True)
	payment_reimbursed_datewise['quantity-reimbursed-inventory'].fillna(0, inplace=True)


	# payment_reimbursed.set_index([''])
	# print(payment_reimbursed.info())

	# payment_reimbursed_datewise.to_excel('payment_reimbursed_datewise.xlsx')
	# sys.exit()

	payment_reimbursed_inventory = payment_reimbursed_datewise.drop(['order-id', 'quantity-refund', 'amount-total', 'quantity-returns', 'quantity-reimbursed-cash', 'quantity-reimbursed-total'], axis=1)
	payment_reimbursed_inventory = payment_reimbursed_inventory[payment_reimbursed_inventory['quantity-reimbursed-inventory'] != 0]
	payment_reimbursed_inventory = payment_reimbursed_inventory.loc[payment_reimbursed_inventory.index.repeat(payment_reimbursed_inventory['quantity-reimbursed-inventory'])].reset_index(drop=True)
	payment_reimbursed_inventory['quantity-reimbursed-inventory'] = 1
	payment_reimbursed_inventory['approval-date'] = payment_reimbursed_inventory['approval-date'].str[0:10]
	payment_reimbursed_inventory['approval-date'] = pd.to_datetime(payment_reimbursed_inventory['approval-date'], format='%Y-%m-%d')

	# print(payment_reimbursed_inventory.info())
	# sys.exit()

	indexer2 = recordlinkage.Index()
	indexer2.block(left_on=['sku', 'fulfillment-center-id'], right_on=['sku', 'fulfillment-center-id'])
	comparisons2 = indexer2.index(payment_reimbursed_inventory, inventory)
	compare2 = recordlinkage.Compare()
	compare2.exact('sku', 'sku', label='sku_match')
	compare2.exact('fulfillment-center-id', 'fulfillment-center-id', label='center_match')
	result2 = compare2.compute(comparisons2, payment_reimbursed_inventory, inventory)
	result_reset2 = result2.reset_index().drop(['sku_match', 'center_match'], axis=1)
	# print(result_reset2)

	result_reset2 = result_reset2.merge(payment_reimbursed_inventory, left_on='level_0', right_index=True)
	result_reset2 = result_reset2.merge(inventory[['Date', 'Quantity']], left_on='level_1', right_index=True)
	result_reset2['date_difference'] = result_reset2['Date'] - result_reset2['approval-date']
	result_reset2 = result_reset2[result_reset2['date_difference'] >= datetime.timedelta(days=0)]
	result_reset2.sort_values(by=['date_difference', 'level_0'], inplace=True)

	returns2 = pd.DataFrame()
	while len(result_reset2) > 0:
		returns2 = returns2.append(result_reset2.iloc[0])
		result_reset2 = result_reset2[(result_reset2['level_0'] != result_reset2['level_0'].iloc[0]) & (result_reset2['level_1'] != result_reset2['level_1'].iloc[0])]

	# returns.to_excel('returns.xlsx')
	# print(returns)

	payment_reimbursed_inventory = payment_reimbursed_inventory.merge(returns2[['level_0', 'Date', 'Quantity', 'date_difference']], left_index=True, right_on='level_0', how='left').drop('level_0', axis=1)
	payment_reimbursed_inventory['quantity-difference'] = np.where(payment_reimbursed_inventory['Quantity'] == 1, 0, 1)
	payment_reimbursed_inventory['amount-difference'] = np.where(payment_reimbursed_inventory['Quantity'] == 1, 0, payment_reimbursed_inventory['total'])

	payment_reimbursed_inventory = payment_reimbursed_inventory.rename(columns={'sku': 'SKU', 'fulfillment-center-id': 'Fulfillment Center', 'Reimbursed': 'Quantity', 'total': 'Amount', 'approval-date': 'Approval Date', 'Date': 'Received Date', 'Quantity': 'Quantity Received', 'date_difference': 'Days Difference', 'quantity-difference': 'Quantity Difference', 'amount-difference': 'Amount Difference'})
	payment_reimbursed_inventory = payment_reimbursed_inventory[['SKU', 'Fulfillment Center', 'Approval Date', 'Quantity', 'Amount', 'Quantity Received', 'Received Date', 'Days Difference', 'Quantity Difference', 'Amount Difference']]
	payment_reimbursed_inventory['Approval Date'] = payment_reimbursed_inventory['Approval Date'].dt.strftime('%m/%d/%Y')
	payment_reimbursed_inventory['Received Date'] = payment_reimbursed_inventory['Received Date'].dt.strftime('%m/%d/%Y')
	payment_reimbursed_inventory.to_excel(data_to_excel, sheet_name='5. Inventory Reimbursement', index=False)

#------------------------------Formatting-------------------------------------------
	workbook = data_to_excel.book
	date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
	number_format = workbook.add_format({'num_format': '#,##0'})
	decimal_format = workbook.add_format({'num_format': '#,##0.00'})
	fail_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
	pass_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
	center_format = workbook.add_format()
	center_format.set_align('center')
	right_format = workbook.add_format()
	right_format.set_align('center')

	sheet1 = data_to_excel.sheets['1. Refunds vs Returns']
	sheet2 = data_to_excel.sheets['2. Returns vs Reimbursements']
	sheet3 = data_to_excel.sheets['3. Cash Reimbursement']
	sheet4 = data_to_excel.sheets['4. Returned to Inventory']
	sheet5 = data_to_excel.sheets['5. Inventory Reimbursement']

	# sheet1.set_column('A:A', 22, center_format)
	sheet1.set_column('A:B', 24, center_format)
	sheet1.set_column('C:C', 24, number_format)
	sheet1.set_column('D:D', 24, decimal_format)
	sheet1.set_column('E:H', 24, number_format)
	sheet1.set_column('I:I', 24, decimal_format)

	sheet2.set_column('A:B', 24, center_format)
	sheet2.set_column('C:C', 24, decimal_format)
	sheet2.set_column('D:D', 24, number_format)
	sheet2.set_column('E:E', 24, decimal_format)
	sheet2.set_column('F:I', 24, number_format)
	sheet2.set_column('J:J', 24, decimal_format)

	sheet3.set_column('A:B', 24, center_format)
	sheet3.set_column('C:C', 24, number_format)
	sheet3.set_column('D:D', 24, decimal_format)
	sheet3.set_column('E:E', 24, number_format)
	sheet3.set_column('F:F', 24, decimal_format)
	sheet3.set_column('G:G', 24, number_format)
	sheet3.set_column('H:H', 24, decimal_format)

	sheet4.set_column('A:C', 24, center_format)
	sheet4.set_column('D:D', 24, number_format)
	sheet4.set_column('E:E', 24, decimal_format)
	sheet4.set_column('F:F', 24, center_format)
	sheet4.set_column('G:I', 24, number_format)
	sheet4.set_column('J:J', 24, decimal_format)

	sheet5.set_column('A:C', 24, center_format)
	sheet5.set_column('D:D', 24, number_format)
	sheet5.set_column('E:E', 24, decimal_format)
	sheet5.set_column('F:F', 24, number_format)
	sheet5.set_column('G:G', 24, center_format)
	sheet5.set_column('H:I', 24, number_format)
	sheet5.set_column('J:J', 24, decimal_format)

	sheet1.conditional_format('H2:I'+str(len(payment_refund_merged_renamed)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet1.conditional_format('H2:I'+str(len(payment_refund_merged_renamed)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet2.conditional_format('I2:J'+str(len(payment_reimbursed_grouped_to_excel)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet2.conditional_format('I2:J'+str(len(payment_reimbursed_grouped_to_excel)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet3.conditional_format('G2:H'+str(len(payment_reimbursed_cash)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet3.conditional_format('G2:H'+str(len(payment_reimbursed_cash)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet4.conditional_format('I2:J'+str(len(payment_inventory_datewise)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet4.conditional_format('I2:J'+str(len(payment_inventory_datewise)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet5.conditional_format('I2:J'+str(len(payment_reimbursed_inventory)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet5.conditional_format('I2:J'+str(len(payment_reimbursed_inventory)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	data_to_excel.save()

	workbook = openpyxl.load_workbook('temp/customer_returns_reco.xlsx')
	sheet1 = workbook['1. Refunds vs Returns']
	sheet2 = workbook['2. Returns vs Reimbursements']
	sheet3 = workbook['3. Cash Reimbursement']
	sheet4 = workbook['4. Returned to Inventory']
	sheet5 = workbook['5. Inventory Reimbursement']

	# print(payment_refund_merged.info())
	# sys.exit()

	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=3).value = payment_refund_merged_renamed['Refund Quantity'].sum()
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=4).value = payment_refund_merged_renamed['Refund Amount'].sum()
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=5).value = payment_refund_merged_renamed['Customer Returns'].sum()			
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=6).value = payment_refund_merged_renamed['Reimbursed'].sum()
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=7).value = payment_refund_merged_renamed['Added to Inventory'].sum()
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=8).value = payment_refund_merged_renamed['Quantity Difference'].sum()
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=9).value = payment_refund_merged_renamed['Amount Difference'].sum()

	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=3).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=6).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=7).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=8).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(payment_refund_merged_renamed)+2, column=9).font = openpyxl.styles.Font(bold=True)

	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=3).value = payment_reimbursed_grouped_to_excel['Refund Amount'].sum()
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=4).value = payment_reimbursed_grouped_to_excel['Reimbursed Quantity'].sum()
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=5).value = payment_reimbursed_grouped_to_excel['Reimbursed Amount'].sum()			
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=6).value = payment_reimbursed_grouped_to_excel['Cash Reimbursement'].sum()
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=7).value = payment_reimbursed_grouped_to_excel['Inventory Reimbursement'].sum()
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=8).value = payment_reimbursed_grouped_to_excel['Total Reimbursement'].sum()
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=9).value = payment_reimbursed_grouped_to_excel['Quantity Difference'].sum()
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=10).value = payment_reimbursed_grouped_to_excel['Amount Difference'].sum()

	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=3).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=6).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=7).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=8).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=9).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(payment_reimbursed_grouped_to_excel)+2, column=10).font = openpyxl.styles.Font(bold=True)

	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=3).value = payment_reimbursed_cash['Quantity Reimbursed'].sum()
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=4).value = payment_reimbursed_cash['Amount Reimbursed'].sum()
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=5).value = payment_reimbursed_cash['Quantity Settled'].sum()			
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=6).value = payment_reimbursed_cash['Amount Settled'].sum()
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=7).value = payment_reimbursed_cash['Quantity Difference'].sum()
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=8).value = payment_reimbursed_cash['Amount Difference'].sum()

	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=3).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=6).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=7).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(payment_reimbursed_cash)+2, column=8).font = openpyxl.styles.Font(bold=True)

	sheet4.cell(row=len(payment_inventory_datewise)+2, column=4).value = payment_inventory_datewise['Return Quantity'].sum()
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=5).value = payment_inventory_datewise['Amount'].sum()
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=7).value = payment_inventory_datewise['Quantity Added'].sum()			
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=9).value = payment_inventory_datewise['Quantity Difference'].sum()
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=10).value = payment_inventory_datewise['Amount Difference'].sum()

	sheet4.cell(row=len(payment_inventory_datewise)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=7).font = openpyxl.styles.Font(bold=True)
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=9).font = openpyxl.styles.Font(bold=True)
	sheet4.cell(row=len(payment_inventory_datewise)+2, column=10).font = openpyxl.styles.Font(bold=True)

	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=4).value = payment_reimbursed_inventory['Quantity'].sum()
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=5).value = payment_reimbursed_inventory['Amount'].sum()
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=6).value = payment_reimbursed_inventory['Quantity Received'].sum()			
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=9).value = payment_reimbursed_inventory['Quantity Difference'].sum()
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=10).value = payment_reimbursed_inventory['Amount Difference'].sum()

	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=6).font = openpyxl.styles.Font(bold=True)
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=9).font = openpyxl.styles.Font(bold=True)
	sheet5.cell(row=len(payment_reimbursed_inventory)+2, column=10).font = openpyxl.styles.Font(bold=True)

	workbook.save('temp/customer_returns_reco.xlsx')

	return

# Clean up unnecessary columns (done)
# Rename columns
# Amounts everywhere (done_)
# Summary
# Single excel (done)
# Formatting
# Tkinter (later)