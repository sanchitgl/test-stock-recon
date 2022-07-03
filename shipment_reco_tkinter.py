import pandas as pd
import os
import sys
from tkinter import *
import customtkinter as ctk
import threading
from tkinter import filedialog
import datetime
import openpyxl

def get_col_widths(dataframe):
	return [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

def show_frame(frame):
	frame.lift()
	# root.iconbitmap('practus_icon.ico')
	root.title('Practus FBA Reconciliation Tool')

def switch_frame(var):
	frame_dict = {1: frame1, 2: frame2}
	frame_dict[var].lift()

def open_file():
	os.system('start EXCEL.EXE C:\\Users\\amitu\\OneDrive\\quantuitix\\projects\\reconcify\\poc\\nmkcdd\\amazon_fba_reco\\output_files\\shipment_reco.xlsx')

def open_folder(path):
	os.system(f'start {os.path.realpath(path)}')

def upload_booking_folder():
	folder_booking_select = filedialog.askdirectory(initialdir='amazon_fba_reco')
	folder_booking_path.set(folder_booking_select)
	folder_booking_name = folder_booking_path.get().split('/')[-1].split('.')[0]
	status1_1_1.configure(text='Completed', fg='green')	
	# status1_2.configure(text=folder_sci_name, fg='#2F5597', font=('MS Reference Sans Serif', 11, 'underline'))#underline=True, underlinefg='#2F5597')
	status1_1_2.configure(text=folder_booking_name, command=lambda:threading.Thread(target=open_folder(folder_booking_path.get())).start())		
	stringvar_booking.set('completed')
	root.update_idletasks()
	enable_reco_button()

def upload_dispatch_folder():
	folder_dispatch_select = filedialog.askdirectory(initialdir='amazon_fba_reco')
	folder_dispatch_path.set(folder_dispatch_select)
	folder_dispatch_name = folder_dispatch_path.get().split('/')[-1].split('.')[0]
	status1_2_1.configure(text='Completed', fg='green')	
	# status2_2.configure(text=folder_inv_name, fg='#2F5597', font=('MS Reference Sans Serif', 11, 'underline'))#underline=True, underlinefg='#2F5597')
	status1_2_2.configure(text=folder_dispatch_name, command=lambda:threading.Thread(target=open_folder(folder_dispatch_path.get())).start())		
	stringvar_dispatch.set('completed')
	root.update_idletasks()
	enable_reco_button()

def upload_inv_folder():
	folder_inv_select = filedialog.askdirectory(initialdir='amazon_fba_reco')
	folder_inv_path.set(folder_inv_select)
	folder_inv_name = folder_inv_path.get().split('/')[-1].split('.')[0]
	status1_3_1.configure(text='Completed', fg='green')	
	# status2_2.configure(text=folder_inv_name, fg='#2F5597', font=('MS Reference Sans Serif', 11, 'underline'))#underline=True, underlinefg='#2F5597')
	status1_3_2.configure(text=folder_inv_name, command=lambda:threading.Thread(target=open_folder(folder_inv_path.get())).start())		
	stringvar_inv.set('completed')
	root.update_idletasks()
	enable_reco_button()

def select_output_folder():
	folder_output_select = filedialog.askdirectory(initialdir='amazon_fba_reco')
	folder_output_path.set(folder_output_select)
	folder_output_name = folder_output_path.get().split('/')[-1]
	status1_4_1.configure(text='Completed', fg='green')
	status1_4_2.configure(text=folder_output_name, command=lambda:threading.Thread(target=open_folder(folder_output_path.get())).start())
	stringvar_output.set('completed')
	root.update_idletasks()
	enable_reco_button()

def enable_reco_button():
	if stringvar_booking.get() == 'completed' and stringvar_inv.get() == 'completed' and stringvar_dispatch.get() == 'completed' and stringvar_output.get() == 'completed':
		# print('Checking StringVar variables')
		button_start.configure(state='normal')
		# print('Reco button enabled')
		root.update_idletasks()
	# else:
		# print('Reco button not enabled')

def run_shipment_reco():
	start_time = datetime.datetime.now()
	try:
		booking_folder = os.listdir(folder_booking_path.get())
	except:
		status1_1_1.configure(text='Not Found', fg='red')
		status1_1_2.configure(text='Not Found', text_color='red')		
		root.update_idletasks()
	else:
		try:
			dispatch_files = os.listdir(folder_dispatch_path.get())
		except:
			status1_2_1.configure(text='Not Found', fg='red')
			status1_2_2.configure(text='Not Found', text_color='red')		
			root.update_idletasks()
		else:

			
			try:
				inventory = pd.read_csv(folder_inv_path.get() + '\\' + 'inventory_ledger.csv')
			except:
				status1_3_1.configure(text='Not Found', fg='red')
				status1_3_2.configure(text='Not Found', text_color='red')			
				root.update_idletasks()
			else:
				try:
					data_to_excel = pd.ExcelWriter(folder_output_path.get() + '\\' + 'shipment_reco.xlsx', engine='xlsxwriter')
				except:
					status1_4_1.configure(text='Not Selected', fg='red')
					status1_4_2.configure(text='Not Selected', text_color='red')
					root.update_idletasks()
				else:
					button_browse1.configure(state='disabled')
					button_browse2.configure(state='disabled')
					button_browse3.configure(state='disabled')
					button_start.configure(state='disabled')
					status1_5_1.configure(text='Running', fg='green')
					root.update_idletasks()
					
					booking = pd.DataFrame()
					for single_folder in booking_folder:
						booking_files = os.listdir(folder_booking_path.get() + '\\' + single_folder)
						for single_booking in booking_files:
							df = pd.read_excel(folder_booking_path.get() + '\\' + single_folder + '\\' + single_booking)
							df['BOOKING DATE'] = single_folder
							df['BOOKING DATE'] = pd.to_datetime(df['BOOKING DATE'], format='%Y_%m_%d')
							df.columns = df.columns.str.strip()
							df.columns = df.columns.str.upper()
							df = df[['FBA ID', 'BOOKING DATE', 'SKU', 'CARTONS', 'QTY']]
							df['FBA ID'] = df['FBA ID'].astype(str)
							df = df[df['FBA ID'].str.len() == 12]
							df['SKU'] = df['SKU'].astype(str)
							df['SKU'] = df['SKU'].replace(regex=True, to_replace=r'(\.0$)', value=r'')
							df['SKU'] = df['SKU'].str.replace('_New', '')
							df['SKU'] = df['SKU'].str.replace('_NEW', '')			
							df['CARTONS'] = df['CARTONS'].astype(int)
							df['QTY'] = df['QTY'].astype(int)
							booking = booking.append(df)

							fba_list = list(set(booking['FBA ID'].to_list()))
							dispatch_filenames = [f + '_ViewTransaction.xlsx' for f in fba_list]

					dispatch = pd.DataFrame()
					for single_dispatch in dispatch_files:
						if single_dispatch in dispatch_filenames:
							wb = openpyxl.load_workbook(folder_dispatch_path.get() + '\\' + single_dispatch)
							ws1 = wb['ViewTransaction']
							fba_id = ws1['R9'].value
							dispatch_date = ws1['AD4'].value
							dispatch_date = dispatch_date.date()

							df = pd.read_excel(folder_dispatch_path.get() + '\\' + single_dispatch, skiprows=45, usecols='C:AS')
							df['FBA ID'] = fba_id
							df['DISPATCH DATE'] = dispatch_date
							df['DISPATCH DATE'] = pd.to_datetime(df['DISPATCH DATE'], format='%Y-%m-%d')
							df.columns = df.columns.str.strip()
							df.columns = df.columns.str.upper()
							df = df[['FBA ID', 'DISPATCH DATE', 'SKU', 'INV QTY']]
							df.dropna(subset=['SKU'], inplace=True)
							remove_cols = [col for col in df.columns if 'Unnamed' in col]
							df.drop(remove_cols, axis=1, inplace=True)
							df['SKU'] = df['SKU'].astype(str)
							df['SKU'] = df['SKU'].replace(regex=True, to_replace=r'(\.0$)', value=r'')
							df['INV QTY'] = df['INV QTY'].astype(int)

							dispatch = dispatch.append(df)

					inventory['Date'] = pd.to_datetime(inventory['Date'], format='%m/%d/%Y')
					inventory['MSKU'] = inventory['MSKU'].astype(str).str[0:12]
					inventory_receipts = inventory[inventory['Event Type'] == 'Receipts'].rename(columns={'Reference ID': 'FBA ID', 'MSKU': 'SKU'})

					inventory_extract = pd.DataFrame()
					for single_fba in fba_list:
						df = inventory_receipts[inventory_receipts['FBA ID'] == single_fba]
						df.columns = df.columns.str.strip()
						df.columns = df.columns.str.upper()
						df = df[['FBA ID', 'DATE', 'SKU', 'QUANTITY']].rename(columns={'DATE': 'RECEIPT DATE'})
						# df = df[['']]
						inventory_extract = inventory_extract.append(df)
					# print(inventory_extract)

					fba_detail = pd.merge(booking, dispatch, on=['FBA ID', 'SKU'], how='outer')
					fba_detail = pd.merge(fba_detail, inventory_extract, on=['FBA ID', 'SKU'], how='outer')
					fba_detail = fba_detail.rename(columns={'CARTONS': 'CARTONS BOOKED', 'INV QTY': 'CARTONS DISPATCHED', 'QTY': 'UNITS BOOKED', 'QUANTITY': 'UNITS RECEIVED'})

					fba_detail['DISPATCH DAYS'] = fba_detail['DISPATCH DATE'] - fba_detail['BOOKING DATE']
					fba_detail['RECEIPT DAYS'] = fba_detail['RECEIPT DATE'] - fba_detail['DISPATCH DATE']

					fba_detail['BOOKING DATE'] = fba_detail['BOOKING DATE'].dt.strftime('%Y-%m-%d')
					fba_detail['DISPATCH DATE'] = fba_detail['DISPATCH DATE'].dt.strftime('%Y-%m-%d')
					fba_detail['RECEIPT DATE'] = fba_detail['RECEIPT DATE'].dt.strftime('%Y-%m-%d')

					fba_detail['BOOKING DATE'].fillna('NOT AVAILABLE', inplace=True)
					fba_detail['DISPATCH DATE'].fillna('NOT AVAILABLE', inplace=True)
					fba_detail['CARTONS BOOKED'].fillna(0, inplace=True)
					fba_detail['CARTONS DISPATCHED'].fillna(0, inplace=True)
					fba_detail['UNITS BOOKED'].fillna(0, inplace=True)
					fba_detail['UNITS RECEIVED'].fillna(0, inplace=True)
					fba_detail['RECEIPT DATE'].fillna('NOT AVAILABLE', inplace=True)
					fba_detail['DISPATCH DAYS'].fillna(datetime.timedelta(days=0), inplace=True)
					fba_detail['RECEIPT DAYS'].fillna(datetime.timedelta(days=0), inplace=True)
					fba_detail.sort_values(by=['FBA ID', 'BOOKING DATE', 'DISPATCH DATE', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED'], inplace=True)

					fba_detail.set_index(['FBA ID', 'BOOKING DATE', 'DISPATCH DATE', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED'], inplace=True)
					# print(fba_detail)
					# fba_detail.to_excel('fba_detail.xlsx')
					# sys.exit()

					fba_skuwise = fba_detail.reset_index()
					fba_skuwise = fba_skuwise.groupby(['FBA ID', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'UNITS BOOKED']).agg({'UNITS RECEIVED': 'sum'}).reset_index()
					fba_skuwise['CARTONS SHORT DISPATCHED'] = fba_skuwise['CARTONS BOOKED'] - fba_skuwise['CARTONS DISPATCHED']
					fba_skuwise['UNITS SHORT RECEIVED'] = fba_skuwise['UNITS BOOKED'] - fba_skuwise['UNITS RECEIVED']
					fba_skuwise.set_index(['FBA ID', 'SKU'], inplace=True)
					fba_skuwise = fba_skuwise[['CARTONS BOOKED', 'CARTONS DISPATCHED', 'CARTONS SHORT DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED', 'UNITS SHORT RECEIVED']]

					fba_exception = fba_skuwise[(fba_skuwise['CARTONS SHORT DISPATCHED'] != 0) | (fba_skuwise['UNITS SHORT RECEIVED'] != 0)]

					sum_positives = lambda x: x[x>0].sum()
					sum_negatives = lambda x: x[x<0].sum() * (-1)
					fba_summary = fba_skuwise.reset_index()
					fba_summary = fba_summary.rename(columns={'CARTONS SHORT DISPATCHED': 'EXCESS CARTONS DISPATCHED', 'UNITS SHORT RECEIVED': 'EXCESS UNITS RECEIVED'})
					fba_summary['SHORT CARTONS DISPATCHED'] = fba_summary['EXCESS CARTONS DISPATCHED']
					fba_summary['SHORT UNITS RECEIVED'] = fba_summary['EXCESS UNITS RECEIVED']
					fba_summary = fba_summary.groupby(['FBA ID']).agg({'SKU': 'count', 'CARTONS BOOKED': 'sum', 'CARTONS DISPATCHED': 'sum', 'UNITS BOOKED': 'sum', 'UNITS RECEIVED': 'sum', 'EXCESS CARTONS DISPATCHED': sum_negatives, 'SHORT CARTONS DISPATCHED': sum_positives, 'EXCESS UNITS RECEIVED': sum_negatives, 'SHORT UNITS RECEIVED': sum_positives}).reset_index()
					fba_summary = fba_summary[['FBA ID', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'EXCESS CARTONS DISPATCHED', 'SHORT CARTONS DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED', 'EXCESS UNITS RECEIVED', 'SHORT UNITS RECEIVED']]

					fba_summary.to_excel(data_to_excel, sheet_name='FBA Summary', index=False)
					fba_exception.to_excel(data_to_excel, sheet_name='FBA Exceptions')
					fba_skuwise.to_excel(data_to_excel, sheet_name='FBA SKU-wise')
					fba_detail.to_excel(data_to_excel, sheet_name='FBA Detail')

					workbook = data_to_excel.book
					date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
					number_format = workbook.add_format({'num_format': '#,##0'})
					fail_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
					pass_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
					center_format = workbook.add_format()
					center_format.set_align('center')
					right_format = workbook.add_format()
					right_format.set_align('center')

					sheet1 = data_to_excel.sheets['FBA Summary']
					sheet2 = data_to_excel.sheets['FBA Exceptions']
					sheet3 = data_to_excel.sheets['FBA SKU-wise']
					sheet4 = data_to_excel.sheets['FBA Detail']

					# sheet1.set_column('A:A', 22, center_format)
					sheet1.set_column('B:J', 22, number_format)
					sheet2.set_column('B:B', 22, center_format)
					sheet2.set_column('C:H', 22, number_format)
					sheet3.set_column('B:B', 22, date_format)
					sheet3.set_column('C:H', 22, number_format)						
					sheet4.set_column('B:D', 22, center_format)
					sheet4.set_column('E:H', 22, number_format)
					sheet4.set_column('I:I', 22, center_format)
					sheet4.set_column('J:K', 22, number_format)

					sheet1.conditional_format('E2:F'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
					sheet1.conditional_format('E2:F'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})
					sheet1.conditional_format('I2:J'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
					sheet1.conditional_format('I2:J'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

					sheet2.conditional_format('E2:E'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
					sheet2.conditional_format('E2:E'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})
					sheet2.conditional_format('H2:H'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
					sheet2.conditional_format('H2:H'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

					sheet3.conditional_format('E2:E'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
					sheet3.conditional_format('E2:E'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})
					sheet3.conditional_format('H2:H'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
					sheet3.conditional_format('H2:H'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

					sheet4.conditional_format('B2:B'+str(len(fba_detail)+1), {'type': 'text', 'criteria': 'containing', 'value': 'NOT AVAILABLE', 'format': fail_format})
					sheet4.conditional_format('C2:C'+str(len(fba_detail)+1), {'type': 'text', 'criteria': 'containing', 'value': 'NOT AVAILABLE', 'format': fail_format})
					sheet4.conditional_format('I2:I'+str(len(fba_detail)+1), {'type': 'text', 'criteria': 'containing', 'value': 'NOT AVAILABLE', 'format': fail_format})
					sheet4.conditional_format('E2:H'+str(len(fba_detail)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': fail_format})

					sheet1.set_column(0, 0, 22)
					sheet2.set_column(0, 0, 22)
					sheet3.set_column(0, 0, 22)
					sheet4.set_column(0, 0, 22)			

					data_to_excel.save()

					workbook = openpyxl.load_workbook(folder_output_path.get() + '\\shipment_reco.xlsx')
					sheet1 = workbook['FBA Summary']
					sheet2 = workbook['FBA Exceptions']
					sheet3 = workbook['FBA SKU-wise']
					sheet4 = workbook['FBA Detail']

					sheet1.cell(row=len(fba_summary)+2, column=2).value = fba_summary['SKU'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=3).value = fba_summary['CARTONS BOOKED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=4).value = fba_summary['CARTONS DISPATCHED'].sum()			
					sheet1.cell(row=len(fba_summary)+2, column=5).value = fba_summary['EXCESS CARTONS DISPATCHED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=6).value = fba_summary['SHORT CARTONS DISPATCHED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=7).value = fba_summary['UNITS BOOKED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=8).value = fba_summary['UNITS RECEIVED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=9).value = fba_summary['EXCESS UNITS RECEIVED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=10).value = fba_summary['SHORT UNITS RECEIVED'].sum()
					sheet1.cell(row=len(fba_summary)+2, column=2).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=3).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=4).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=5).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=6).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=7).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=8).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=9).font = openpyxl.styles.Font(bold=True)
					sheet1.cell(row=len(fba_summary)+2, column=10).font = openpyxl.styles.Font(bold=True)

					sheet2.cell(row=len(fba_exception)+2, column=5).value = fba_exception['CARTONS SHORT DISPATCHED'].sum()
					sheet2.cell(row=len(fba_exception)+2, column=8).value = fba_exception['UNITS SHORT RECEIVED'].sum()
					sheet2.cell(row=len(fba_exception)+2, column=5).font = openpyxl.styles.Font(bold=True)
					sheet2.cell(row=len(fba_exception)+2, column=8).font = openpyxl.styles.Font(bold=True)

					sheet3.cell(row=len(fba_skuwise)+2, column=3).value = fba_skuwise['CARTONS BOOKED'].sum()
					sheet3.cell(row=len(fba_skuwise)+2, column=4).value = fba_skuwise['CARTONS DISPATCHED'].sum()
					sheet3.cell(row=len(fba_skuwise)+2, column=5).value = fba_skuwise['CARTONS SHORT DISPATCHED'].sum()			
					sheet3.cell(row=len(fba_skuwise)+2, column=6).value = fba_skuwise['UNITS BOOKED'].sum()
					sheet3.cell(row=len(fba_skuwise)+2, column=7).value = fba_skuwise['UNITS RECEIVED'].sum()
					sheet3.cell(row=len(fba_skuwise)+2, column=8).value = fba_skuwise['UNITS SHORT RECEIVED'].sum()
					sheet3.cell(row=len(fba_skuwise)+2, column=3).font = openpyxl.styles.Font(bold=True)
					sheet3.cell(row=len(fba_skuwise)+2, column=4).font = openpyxl.styles.Font(bold=True)
					sheet3.cell(row=len(fba_skuwise)+2, column=5).font = openpyxl.styles.Font(bold=True)
					sheet3.cell(row=len(fba_skuwise)+2, column=6).font = openpyxl.styles.Font(bold=True)
					sheet3.cell(row=len(fba_skuwise)+2, column=7).font = openpyxl.styles.Font(bold=True)
					sheet3.cell(row=len(fba_skuwise)+2, column=8).font = openpyxl.styles.Font(bold=True)

					for c in range(1, 11):
						sheet1.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
					# 	sheet1.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

					for c in range(1, 9):
						sheet2.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
					# 	sheet2.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

					for c in range(1, 9):
						sheet3.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
					# 	sheet3.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

					for c in range(1, 12):
						sheet4.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
					# 	sheet4.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

					for c in ('A', 'B'):
						for r in range(2, len(fba_exception)+2):
							sheet2[c+str(r)].font = openpyxl.styles.Font(bold=False)

					for c in ('A', 'B'):
						for r in range(2, len(fba_skuwise)+2):
							sheet3[c+str(r)].font = openpyxl.styles.Font(bold=False)

					for c in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
						for r in range(2, len(fba_detail)+2):
							sheet4[c+str(r)].font = openpyxl.styles.Font(bold=False)

					border = openpyxl.styles.borders.Side(style = None, color = '00000000', border_style = 'thin')
					border_format = openpyxl.styles.borders.Border(left = border, right = border, bottom = border, top = border)

					for c in ('C', 'D', 'E', 'F', 'G', 'H'):
						for r in range(2, len(fba_exception)+3):
							sheet2[c+str(r)].border = border_format

					for c in ('C', 'D', 'E', 'F', 'G', 'H'):
						for r in range(2, len(fba_skuwise)+3):
							sheet3[c+str(r)].border = border_format

					for c in ('I', 'J', 'K'):
						for r in range(2, len(fba_detail)+2):
							sheet4[c+str(r)].border = border_format

					workbook.save(folder_output_path.get() + '\\shipment_reco.xlsx')

					end_time = datetime.datetime.now()
					processing_time = end_time - start_time
					processing_seconds = round(processing_time.total_seconds(),4)
					status1_5_1.configure(text='Completed in '+str(processing_seconds)+' s', fg='green')
					status1_5_2.configure(text='shipment_reco.xlsx', command=lambda:threading.Thread(target=open_file).start())		
					button_email.configure(state='normal')
					root.update_idletasks()

def reset_all():
	button_browse1.configure(state='normal')
	button_browse2.configure(state='normal')
	button_browse3.configure(state='normal')
	button_start.configure(state='disabled')
	button_email.configure(state='disabled')
	
	status1_1_1.configure(text='Not Started', fg='orange')
	status1_2_1.configure(text='Not Started', fg='orange')
	status1_3_1.configure(text='Not Started', fg='orange')
	status1_4_1.configure(text='Not Started', fg='orange')
	status1_5_1.configure(text='Not Started', fg='orange')
	
	status1_1_2.configure(text='None', command=lambda:threading.Thread(target=show_frame(frame1)).start())
	status1_2_2.configure(text='None', command=lambda:threading.Thread(target=show_frame(frame1)).start())
	status1_3_2.configure(text='None', command=lambda:threading.Thread(target=show_frame(frame1)).start())
	status1_4_2.configure(text='None', command=lambda:threading.Thread(target=show_frame(frame1)).start())
	status1_5_2.configure(text='None', command=lambda:threading.Thread(target=show_frame(frame1)).start())
	
	stringvar_booking.set('not completed')
	stringvar_inv.set('not completed')
	stringvar_dispatch.set('not completed')
	stringvar_output.set('not completed')

ctk.set_appearance_mode('system')  # Modes: system (default), light, dark
ctk.set_default_color_theme('blue')  # Themes: blue (default), dark-blue, green

root = ctk.CTk()
root.title('')
root.geometry('900x520')
root.configure(background='white')

root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)

frame0 = ctk.CTkFrame(root, corner_radius= 10, fg_color='#F0F0F0', bg_color='white')
frame1 = ctk.CTkFrame(root, corner_radius= 10, fg_color='#F0F0F0', bg_color='white')
frame2 = Frame(root, bg='white')
frame3 = Frame(root, bg='white')
frame4 = Frame(root, bg='white')

for frame in (frame0, frame1, frame2, frame3):
    frame.grid(row=0, column=0, sticky='nsew')

#----------------------------------------Frame 0--------------------------------------------

#practus_logo = PhotoImage(file='practus_fba2_480.png')
#ctk.CTkLabel(frame0,image=practus_logo).place(x=35, y=35)

#reconcify_logo = PhotoImage(file='powered_by_resized.png')
#ctk.CTkLabel(frame0,image=reconcify_logo).place(x=120, y=450)

#background_img = PhotoImage(file='amazon_fba_250.png')
#ctk.CTkLabel(frame0,image=background_img).place(x=100, y=160)

# label_1 = ctk.CTkLabel(master=frame0, text="Fulfilled-by-Amazon", text_color='black', text_font=('MS Reference Sans Serif', 13), anchor='e')
# label_1.place(x=230, y=192)

# label_2 = ctk.CTkLabel(master=frame0, text="Reconciliation Tool", text_color='black', text_font=('MS Reference Sans Serif', 13), anchor='e')
# label_2.place(x=230, y=215)

# label_3 = ctk.CTkLabel(master=frame0, text="P O W E R E D   B Y", text_color='black', text_font=('MS Reference Sans Serif', 6), anchor='e')
# label_3.place(x=150, y=340)

# radio_var = IntVar(0)
# radiobutton_1 = ctk.CTkRadioButton(master=frame0, text='FBA Shipments', variable= radio_var, value=1, hover=True, text_font=('MS Reference Sans Serif', 11))
# radiobutton_1.place(x=580, y=120)

# radiobutton_2 = ctk.CTkRadioButton(master=frame0, text='Sales Orders', variable= radio_var, value=2, hover=True, text_font=('MS Reference Sans Serif', 11))
# radiobutton_2.place(x=580, y=170)

# radiobutton_3 = ctk.CTkRadioButton(master=frame0, text='Sales Returns', variable= radio_var, value=3, hover=True, text_font=('MS Reference Sans Serif', 11))
# radiobutton_3.place(x=580, y=220)

# radiobutton_4 = ctk.CTkRadioButton(master=frame0, text='Customer Damages', variable= radio_var, value=4, hover=True, text_font=('MS Reference Sans Serif', 11))
# radiobutton_4.place(x=580, y=270)

# radiobutton_5 = ctk.CTkRadioButton(master=frame0, text='Warehouse Damages', variable= radio_var, value=5, hover=True, text_font=('MS Reference Sans Serif', 11))
# radiobutton_5.place(x=580, y=320)

# radiobutton_6 = ctk.CTkRadioButton(master=frame0, text='Lost or Misplaced', variable= radio_var, value=6, hover=True, text_font=('MS Reference Sans Serif', 11))
# radiobutton_6.place(x=580, y=370)

button_1 = ctk.CTkButton(master=frame0, width=200, height=36, borderwidth=0, corner_radius=0, text='FBA Shipments', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 11), command=lambda:threading.Thread(target=switch_frame(1)).start()) 
button_1.place(x=570, y=150)

button_2 = ctk.CTkButton(master=frame0, width=200, height=36, borderwidth=0, corner_radius=0, text='Sales Orders', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 11), command=lambda:threading.Thread(target=switch_frame(2)).start())                                 
button_2.place(x=570, y=200)

button_3 = ctk.CTkButton(master=frame0, width=200, height=36, borderwidth=0, corner_radius=0, text='Sales Returns', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 11), command=lambda:threading.Thread(target=switch_frame(3)).start())
button_3.place(x=570, y=250)

button_4 = ctk.CTkButton(master=frame0, width=200, height=36, borderwidth=0, corner_radius=0, text='Customer Damages', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 11), command=lambda:threading.Thread(target=switch_frame(4)).start())                                 
button_4.place(x=570, y=300)

button_5 = ctk.CTkButton(master=frame0, width=200, height=36, borderwidth=0, corner_radius=0, text='Warehouse Damages', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 11), command=lambda:threading.Thread(target=switch_frame(5)).start())                                 
button_5.place(x=570, y=350)

button_6 = ctk.CTkButton(master=frame0, width=200, height=36, borderwidth=0, corner_radius=0, text='Lost or Misplaced', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 11), command=lambda:threading.Thread(target=switch_frame(6)).start())                                 
button_6.place(x=570, y=400)


#--------------------------------------Frame 1-------------------------------------
# practus_logo1 = PhotoImage(file='practus_logo_resized.png')
# Label(frame1,image=practus_logo1).place(x=20, y=10)

# reconcify_logo1 = PhotoImage(file='reconcify_logo_200.png')
# Label(frame1,image=reconcify_logo1).place(x=680, y=15)

bot_name = Label(frame1, text='FBA Shipments Reconciliation', font=('MS Reference Sans Serif', 13), fg='black')
bot_name.place(x=320, y=100)

step1 = Label(frame1, text='Upload Shipment Instructions', font=('MS Reference Sans Serif', 11), fg='#2F5597')
step1.place(x=50, y=160)

step2 = Label(frame1, text='Upload Warehouse Reports', font=('MS Reference Sans Serif', 11), fg='#2F5597')
step2.place(x=50, y=210)

step3 = Label(frame1, text='Upload inventory Ledger', font=('MS Reference Sans Serif', 11), fg='#2F5597')
step3.place(x=50, y=260)

step4 = Label(frame1, text='Select Output Folder', font=('MS Reference Sans Serif', 11), fg='#2F5597')
step4.place(x=50, y=310)

step5 = Label(frame1, text='Start Reconciliation', font=('MS Reference Sans Serif', 11), fg='#2F5597')
step5.place(x=50, y=360)

folder_booking_path = StringVar()
folder_inv_path = StringVar()
folder_dispatch_path = StringVar()
folder_output_path = StringVar()

stringvar_booking = StringVar()
stringvar_inv = StringVar()
stringvar_dispatch = StringVar()
stringvar_output = StringVar()

stringvar_booking.set('not completed')
stringvar_inv.set('not completed')
stringvar_dispatch.set('not completed')
stringvar_output.set('not completed')

button_browse1 = ctk.CTkButton(master=frame1, width=100, height=30, borderwidth=0, corner_radius=0, text='Browse', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), command=lambda:threading.Thread(target=upload_booking_folder).start())                             
button_browse1.place(x=320, y=160)

button_browse2 = ctk.CTkButton(master=frame1, width=100, height=30, borderwidth=0, corner_radius=0, text='Browse', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), command=lambda:threading.Thread(target=upload_dispatch_folder).start())                             
button_browse2.place(x=320, y=210)

button_browse3 = ctk.CTkButton(master=frame1, width=100, height=30, borderwidth=0, corner_radius=0, text='Browse', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), command=lambda:threading.Thread(target=upload_inv_folder).start())                             
button_browse3.place(x=320, y=260)

button_select = ctk.CTkButton(master=frame1, width=100, height=30, borderwidth=0, corner_radius=0, text='Select', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), command=lambda:threading.Thread(target=select_output_folder).start())                             
button_select.place(x=320, y=310)

button_start = ctk.CTkButton(master=frame1, width=100, height=30, borderwidth=0, corner_radius=0, text='Start', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), state='disabled', command=lambda:threading.Thread(target=run_shipment_reco).start())
button_start.place(x=320, y=360)

status1_1_1 = Label(frame1, text='Not Started', font=('MS Reference Sans Serif', 11), fg='orange')
status1_1_1.place(x=500, y=160)

status1_2_1 = Label(frame1, text='Not Started', font=('MS Reference Sans Serif', 11), fg='orange')
status1_2_1.place(x=500, y=210)

status1_3_1 = Label(frame1, text='Not Started', font=('MS Reference Sans Serif', 11), fg='orange')
status1_3_1.place(x=500, y=260)

status1_4_1 = Label(frame1, text='Not Started', font=('MS Reference Sans Serif', 11), fg='orange')
status1_4_1.place(x=500, y=310)

status1_5_1 = Label(frame1, text='Not Started', font=('MS Reference Sans Serif', 11), fg='orange')
status1_5_1.place(x=500, y=360)

status1_1_2 = ctk.CTkButton(master=frame1, width=100, height=10, borderwidth=0, corner_radius=8, text='None', text_font=('MS Reference Sans Serif', 11, 'underline'), text_color='blue', hover=False, fg_color='#F0F0F0', command=lambda:threading.Thread(target=show_frame(frame1)).start())		
status1_1_2.place(x=710, y=160)

status1_2_2 = ctk.CTkButton(master=frame1, width=100, height=10, borderwidth=0, corner_radius=8, text='None', text_font=('MS Reference Sans Serif', 11, 'underline'), text_color='blue', hover=False, fg_color='#F0F0F0', command=lambda:threading.Thread(target=show_frame(frame1)).start())		
status1_2_2.place(x=710, y=210)

status1_3_2 = ctk.CTkButton(master=frame1, width=100, height=10, borderwidth=0, corner_radius=8, text='None', text_font=('MS Reference Sans Serif', 11, 'underline'), text_color='blue', hover=False, fg_color='#F0F0F0', command=lambda:threading.Thread(target=show_frame(frame1)).start())		
status1_3_2.place(x=710, y=260)

status1_4_2 = ctk.CTkButton(master=frame1, width=100, height=10, borderwidth=0, corner_radius=8, text='None', text_font=('MS Reference Sans Serif', 11, 'underline'), text_color='blue', hover=False, fg_color='#F0F0F0', command=lambda:threading.Thread(target=show_frame(frame1)).start())		
status1_4_2.place(x=710, y=310)

status1_5_2 = ctk.CTkButton(master=frame1, width=100, height=10, borderwidth=0, corner_radius=8, text='None', text_font=('MS Reference Sans Serif', 11, 'underline'), text_color='blue', hover=False, fg_color='#F0F0F0', command=lambda:threading.Thread(target=show_frame(frame1)).start())		
status1_5_2.place(x=710, y=360)

button_email = ctk.CTkButton(master=frame1, width=120, height=33, borderwidth=0, corner_radius=0, text='Email Reco', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), state='disabled')#, command=lambda:threading.Thread(target=run_shipment_reco).start())
button_email.place(x=260, y=440)

button_reset = ctk.CTkButton(master=frame1, width=120, height=33, borderwidth=0, corner_radius=0, text='Reset', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), command=lambda:threading.Thread(target=reset_all).start())
button_reset.place(x=390, y=440)

button_back = ctk.CTkButton(master=frame1, width=120, height=33, borderwidth=0, corner_radius=0, text='Back', fg_color='#FFDF8C', hover_color='#FFC446', text_font=('MS Reference Sans Serif', 10), command=lambda:threading.Thread(target=show_frame(frame0)).start())
button_back.place(x=520, y=440)

show_frame(frame0)

root.mainloop()

# Output path
# Dispatch path
# Formatting