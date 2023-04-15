import pandas as pd
import os
import sys
import numpy as np
import openpyxl

def reconcile(packing_lists, inventory_ledger):
	# packing_list_path = 'C:\\Users\\amitu\\OneDrive\\quantuitix\\projects\\reconcify\\poc\\nmkcdd\\japan\\input_files\\packing_list'
	# packing_list_files = os.listdir(packing_list_path)
	# # print(packing_list_files)

	packing_list = pd.DataFrame()
	for file in packing_lists:
		wb = openpyxl.load_workbook(file, data_only=True)
		ws = wb['Packing List']
		for cell in ws['B']:
		# for cell in row:
			if cell.value == 'Carton From':
				row = cell.row
			if cell.value == 'BL/FCR/AWB DATE':
				row = cell.row
				dispatch_date = ws['C'+str(row)].value
				# print(dispatch_date)
		# sys.exit()

				# print(row)
		# sys.exit()

		df = pd.read_excel(file, sheet_name='Packing List', skiprows=row-1, usecols='B:M')
		df['UPC Number'] = df['UPC Number'].astype(str)
		# df = df[(df['Carton From'] != 'SUB TOTAL') & (df['Carton From'] != 'SUB TOTAL') (df['Carton From'].str[0:2] != 'Po')]
		df.dropna(subset=['FBA Id /  Carton Id'], inplace=True)
		df['BL Date'] = dispatch_date
		packing_list = packing_list.append(df)
		
	packing_list['FBA ID'] = packing_list['FBA Id /  Carton Id'].str[:12]
	packing_list.reset_index(drop=True, inplace=True)
	# print(packing_list)

	# inventory_ledger_path = 'C:\\Users\\amitu\\OneDrive\\quantuitix\\projects\\reconcify\\poc\\nmkcdd\\japan\\input_files\\inventory_ledger'
	# inventory_ledger_file = os.listdir(inventory_ledger_path)
	inventory_ledger = pd.read_csv(inventory_ledger)
	inventory_receipts = inventory_ledger[inventory_ledger['Event Type'] == 'Receipts']
	inventory_receipts['MSKU'] = inventory_receipts['MSKU'].astype(str)
	# print(inventory_receipts)

	# for fba_id in list(set(packing_list['FBA_ID'].to_list())):
	# 	packing_list_extract = packing_list[packing_list['FBA_ID'] == fba_id]
	# 	inventory_receipts_extract = inventory_receipts[inventory_receipts['Reference ID'] == fba_id]
	# 	print(packing_list_extract)
	# 	print(inventory_receipts_extract)

	packing_list_grouped = packing_list.groupby(['FBA ID', 'UPC Number']).agg({'Total Qty': 'sum'}).reset_index().rename(columns={'Total Qty': 'Qty: Packing List', 'UPC Number': 'SKU ID'})
	inventory_receipts_grouped = inventory_receipts.groupby(['Reference ID', 'MSKU']).agg({'Quantity': 'sum'}).reset_index().rename(columns={'Reference ID': 'FBA ID', 'MSKU': 'SKU ID', 'Quantity': 'Qty: Inv Ledger'})
	# print(packing_list_grouped)
	# sys.exit()
	# print(inventory_receipts_grouped[''])

	summary = pd.merge(packing_list_grouped, inventory_receipts_grouped, on=['FBA ID', 'SKU ID'], how='left')
	summary['Qty: Packing List'].fillna(0, inplace=True)
	summary['Qty: Inv Ledger'].fillna(0, inplace=True)
	summary['Excess/(Short) Received'] = summary['Qty: Inv Ledger'] - summary['Qty: Packing List']
	# summary['SKU ID'] = summary['SKU ID'].astype(str)
	summary.set_index(['FBA ID', 'SKU ID'], inplace=True)
	# print(summary)

	packing_list_grouped2 = packing_list.groupby(['FBA ID', 'UPC Number', 'BL Date']).agg({'Total Qty': 'sum'}).reset_index().rename(columns={'Total Qty': 'Qty: Packing List', 'UPC Number': 'SKU ID'})
	inventory_receipts_grouped2 = inventory_receipts.groupby(['Reference ID', 'MSKU', 'Date']).agg({'Quantity': 'sum'}).reset_index().rename(columns={'Reference ID': 'FBA ID', 'MSKU': 'SKU ID', 'Quantity': 'Qty: Inv Ledger', 'Date': 'Receipt Date'})
	inventory_receipts_grouped2['Receipt Date'] = pd.to_datetime(inventory_receipts_grouped2['Receipt Date'], format='%m/%d/%Y')

	detailed = pd.merge(packing_list_grouped2, inventory_receipts_grouped2, on=['FBA ID', 'SKU ID'], how='left')
	detailed['Transit Days'] = detailed['Receipt Date'] - detailed['BL Date']
	detailed['BL Date'] = detailed['BL Date'].dt.strftime('%Y-%m-%d')
	detailed['Receipt Date'] = detailed['Receipt Date'].dt.strftime('%Y-%m-%d')
	detailed['Qty: Packing List'].fillna(0, inplace=True)
	detailed['Qty: Inv Ledger'].fillna(0, inplace=True)
	detailed['BL Date'].fillna('-', inplace=True)
	detailed['Receipt Date'].fillna('-', inplace=True)
	# detailed['Transit Days'].fillna('-', inplace=True)
	# detailed['SKU ID'] = detailed['SKU ID'].astype(str)
	detailed.set_index(['FBA ID', 'SKU ID', 'Qty: Packing List', 'BL Date'], inplace=True)
	detailed = detailed[['Qty: Inv Ledger', 'Receipt Date', 'Transit Days']]
	# detailed['Qty: Difference'] = detailed['Qty: Packing List'] - detailed['Qty: Inv Ledger']
	# print(detailed)
	# detailed = pd.merge(pa)

	writer = pd.ExcelWriter('temp/fba_reco_japan.xlsx')
	summary.to_excel(writer, sheet_name='SKU-wise')
	detailed.to_excel(writer, sheet_name='SKU-wise Date-wise')

	workbook = writer.book
	fail_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
	pass_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
	center_format = workbook.add_format()
	center_format.set_align('center')
	# right_format = workbook.add_format()
	# right_format.set_align('center')

	sheet1 = writer.sheets['SKU-wise']
	sheet2 = writer.sheets['SKU-wise Date-wise']

	# sheet1.set_column('A:A', 22, center_format)
	sheet1.set_column('A:E', 22)
	sheet2.set_column('A:G', 22)
	sheet2.set_column('F:F', 22, center_format)

	sheet1.conditional_format('E2:E'+str(len(summary)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet1.conditional_format('E2:E'+str(len(summary)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	writer.save()

	workbook = openpyxl.load_workbook('temp/fba_reco_japan.xlsx')
	sheet1 = workbook['SKU-wise']
	sheet2 = workbook['SKU-wise Date-wise']

	for c in ('A', 'B'):
		for r in range(2, len(summary)+2):
			sheet1[c+str(r)].font = openpyxl.styles.Font(bold=False)

	for c in ('A', 'B', 'C', 'D'):
		for r in range(2, len(detailed)+2):
			sheet2[c+str(r)].font = openpyxl.styles.Font(bold=False)

	border = openpyxl.styles.borders.Side(style = None, color = '00000000', border_style = 'thin')
	border_format = openpyxl.styles.borders.Border(left = border, right = border, bottom = border, top = border)

	for c in ('C', 'D', 'E'):
		for r in range(2, len(summary)+2):
			sheet1[c+str(r)].border = border_format

	for c in ('E', 'F', 'G'):
		for r in range(2, len(detailed)+2):
			sheet2[c+str(r)].border = border_format

	workbook.save('temp/fba_reco_japan.xlsx')

	return